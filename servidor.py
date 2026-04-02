"""
servidor.py — API Flask para procesar Indicadores
Deploy en Render.com (gratis)
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os, io, glob, traceback, tempfile
import pandas as pd
import numpy as np
import holidays
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font

app = Flask(__name__)
CORS(app)


# =============================================================================
#  Ping
# =============================================================================
@app.route("/ping")
def ping():
    return jsonify({"status": "ok"})


# =============================================================================
#  Endpoint principal
# =============================================================================
@app.route("/correr-indicadores", methods=["POST"])
def correr_indicadores():
    tmp = tempfile.mkdtemp()
    try:
        # ── 1. Leer parámetros ──────────────────────────────────────────────
        mes_  = int(request.form.get("mes"))
        anio_ = int(request.form.get("anio"))
        mes   = f"{mes_:02d}"
        anio  = str(anio_)

        # ── 2. Guardar archivos subidos en carpeta temporal ─────────────────
        archivos_recibidos = {}
        for key, f in request.files.items():
            ruta = os.path.join(tmp, f.filename)
            f.save(ruta)
            archivos_recibidos[key] = ruta

        # Verificar archivos obligatorios
        requeridos = [
            "demanda", "openpass", "pmhab",
            "dmt", "goef", "rutap", "indicadores"
        ]
        faltantes = [r for r in requeridos if r not in archivos_recibidos]
        if faltantes:
            return jsonify({"error": f"Faltan archivos: {', '.join(faltantes)}"}), 400

        # ── 3. Fechas y días hábiles ────────────────────────────────────────
        fecha_i = pd.to_datetime(f'{anio}-{mes}-01')
        fecha_f = fecha_i + pd.offsets.MonthBegin(1)

        fechas_mes  = pd.date_range(start=fecha_i, end=fecha_f - pd.Timedelta(days=1))
        df_dias     = pd.DataFrame({'Fecha': fechas_mes})
        dias_map    = {0:'Lunes',1:'Martes',2:'Miercoles',3:'Jueves',
                       4:'Viernes',5:'Sabado',6:'Domingo'}
        df_dias['Dia']  = df_dias['Fecha'].dt.dayofweek.map(dias_map)
        feriados_ar     = holidays.Argentina(years=[anio_])
        df_dias['Tipo'] = df_dias['Fecha'].apply(
            lambda x: 'DNH' if (x.weekday() >= 5 or x in feriados_ar) else 'DH'
        )
        df_dias['Mes'] = df_dias['Fecha'].dt.month.astype(str).str.zfill(2)
        cuenta_DH  = df_dias[(df_dias['Mes'] == mes) & (df_dias['Tipo'] == 'DH')].shape[0]
        cuenta_DSH = df_dias[(df_dias['Mes'] == mes) & (df_dias['Tipo'] == 'DSH')].shape[0]
        cuenta_DNH = df_dias[(df_dias['Mes'] == mes) & (df_dias['Tipo'] == 'DNH')].shape[0]
        dias_habiles = cuenta_DH + (cuenta_DSH / 2) + (cuenta_DNH / 4)

        meses_dict = {
            '01':'Enero','02':'Febrero','03':'Marzo','04':'Abril',
            '05':'Mayo','06':'Junio','07':'Julio','08':'Agosto',
            '09':'Septiembre','10':'Octubre','11':'Noviembre','12':'Diciembre'
        }

        # ── 4. DEMANDA (output de script anterior) ──────────────────────────
        df_demanda_final_ = pd.read_excel(archivos_recibidos["demanda"])

        # ── 5. OFERTA (GOEF - Velocidad Comercial) ──────────────────────────
        df_goef = pd.read_csv(archivos_recibidos["goef"], sep=';', usecols=range(6))
        df_goef['VC'] = (
            df_goef['VC'].astype(str)
            .str.replace(',', '.', regex=False)
            .astype(float).round(2)
        )
        df_goef = df_goef.rename(columns={'VC': 'VC_goef'})
        df_goef['TIEMPO'] = (
            pd.to_timedelta(df_goef['TIEMPO'].astype(str))
            .dt.total_seconds().div(3600).round(2)
        )
        df_goef['KM'] = pd.to_numeric(df_goef['KM'], errors='coerce')
        df_goef = df_goef.rename(columns={
            'LINEA COMERCIAL': 'Línea',
            'TIEMPO': 'Tiempo de marcha mensual (en hs)',
            'KM': 'Km mensuales'
        })

        # ── 6. MERGE Demanda + Oferta ───────────────────────────────────────
        df_demanda_oferta = df_demanda_final_.merge(
            df_goef, on='IDLINEA', how='left', suffixes=('', '_goef')
        )
        if 'GT' in df_demanda_oferta.columns:
            df_demanda_oferta = df_demanda_oferta.drop(columns=['GT'])

        empresa_reciente = (
            df_demanda_oferta
            .sort_values('Fecha_ELR', ascending=False, na_position='last')
            .drop_duplicates('Línea')[['Línea', 'Empresa']]
        ) if 'Fecha_ELR' in df_demanda_oferta.columns else (
            df_demanda_oferta.drop_duplicates('Línea')[['Línea', 'Empresa']]
        )

        agg_cols = {
            'JURISDICCION': 'min',
            'TRX SUBE física': 'sum',
            'TRX app SUBE': 'sum',
            'TRX QR': 'sum',
            'TRX tarjetas bancarias (física o virtual)': 'sum',
            'TRX con boleto escolar primario': 'sum',
            'TRX con boleto escolar secundario': 'sum',
            'TRX con ATS': 'sum',
            'TRX a partir de 2da integración (ITG)': 'sum',
            'TRX multipago (QRs + tarjetas)': 'sum',
            'TRX SUBE (física + app)': 'sum',
            'TRX total NSSA': 'sum',
            'Max Veh Calle': 'max',
            'Km mensuales': 'sum',
            'Tiempo de marcha mensual (en hs)': 'sum',
            'VC_goef': 'mean'
        }
        # Solo agregar columnas que existen
        agg_cols = {k: v for k, v in agg_cols.items() if k in df_demanda_oferta.columns}
        df_demanda_oferta = df_demanda_oferta.groupby('Línea', as_index=False).agg(agg_cols)
        df_demanda_oferta = df_demanda_oferta.merge(empresa_reciente, on='Línea', how='left')
        df_demanda_oferta['VC'] = (
            df_demanda_oferta['Km mensuales'] /
            df_demanda_oferta['Tiempo de marcha mensual (en hs)']
        ).round(2)
        df_demanda_oferta['prom TRX dia habil'] = df_demanda_oferta['TRX total NSSA'] / dias_habiles

        # ── 7. OPEN PASS ────────────────────────────────────────────────────
        df_op = pd.read_csv(archivos_recibidos["openpass"])
        df_op['fecha_tap'] = pd.to_datetime(df_op['fecha_tap']).dt.strftime('%Y-%m')
        df_op = df_op.rename(columns={'fecha_tap': 'YYYY-MM'})
        df_op_mensual = df_op.groupby('YYYY-MM', as_index=False).sum()

        fecha_objetivo = f'{anio}-{mes}'
        df_mes = df_op_mensual[df_op_mensual['YYYY-MM'] == fecha_objetivo]
        df_long = df_mes.melt(
            id_vars=['YYYY-MM'],
            var_name='Línea',
            value_name='TRX OP tarjetas bancarias (física o virtual)'
        )[['Línea', 'TRX OP tarjetas bancarias (física o virtual)']]
        df_long = df_long[df_long['Línea'] != 'sin_nro_linea']
        df_long['Línea'] = pd.to_numeric(df_long['Línea'], errors='coerce')

        df_demanda_oferta_op = df_demanda_oferta.merge(df_long, on='Línea', how='left')
        df_demanda_oferta_op['TRX OP tarjetas bancarias (física o virtual)'] = pd.to_numeric(
            df_demanda_oferta_op['TRX OP tarjetas bancarias (física o virtual)'].fillna(0),
            errors='coerce'
        )
        df_demanda_oferta_op['TRX total mes'] = (
            df_demanda_oferta_op['TRX OP tarjetas bancarias (física o virtual)'] +
            df_demanda_oferta_op['TRX total NSSA']
        )
        df_demanda_oferta_op['prom TRX dia habil'] = df_demanda_oferta_op['TRX total mes'] / dias_habiles
        df_demanda_oferta_op['Km día hábil']        = df_demanda_oferta_op['Km mensuales'] / dias_habiles
        df_demanda_oferta_op = df_demanda_oferta_op.sort_values('Línea')

        # ── 8. PARQUE MÓVIL (RUTAP) ─────────────────────────────────────────
        ruta_rutap = archivos_recibidos["rutap"]
        df_parque  = pd.read_excel(ruta_rutap)
        df_parque['Tipo_Motor'] = df_parque['Tipo_Motor'].str.strip()

        df_motores = (
            df_parque.groupby('Linea')['Tipo_Motor']
            .value_counts(normalize=True)
            .unstack(fill_value=0) * 100
        ).reset_index()
        df_motores.columns.name = None
        rename_motores = {
            'E3': 'Motor Euro3', 'E5': 'Motor Euro5',
            'Electrico': 'Motor Eléctrico', 'GNC': 'Motor a GNC'
        }
        df_motores.rename(columns={k: v for k, v in rename_motores.items()
                                    if k in df_motores.columns}, inplace=True)

        df_parque_final_ = df_parque.groupby('Linea')['Año_Modelo'].mean().reset_index()
        df_parque_final_['Antiguedad media de flota'] = (anio_ - df_parque_final_['Año_Modelo']).round(2)
        df_parque_final_ = df_parque_final_.drop(columns=['Año_Modelo'])
        df_parque_final  = pd.merge(df_parque_final_, df_motores, on='Linea', how='left')
        df_parque_final  = df_parque_final.rename(columns={'Linea': 'Línea'})
        df_parque_final.iloc[:, 0] = df_parque_final.iloc[:, 0].astype(int)
        df_parque_final.iloc[:, 1:] = df_parque_final.iloc[:, 1:].round(2)
        cols_pct = df_parque_final.columns[2:]
        df_parque_final[cols_pct] = df_parque_final[cols_pct].apply(lambda col: col.map(lambda x: f"{x:.2f}%"))
        df_parque_final['Línea'] = df_parque_final['Línea'].astype(int).replace(0, 1057)

        # ── 9. PARQUE HABILITADO ────────────────────────────────────────────
        df_pmhab = pd.read_excel(archivos_recibidos["pmhab"])[['Línea', 'PM Hab']]
        df_parque_final_2 = df_pmhab.merge(df_parque_final, on='Línea', how='left')

        # ── 10. Merge Demanda + Parque ──────────────────────────────────────
        df_dem_op_parque = df_demanda_oferta_op.merge(df_parque_final_2, on='Línea', how='left')
        df_dem_op_parque = df_dem_op_parque.sort_values('Línea')

        # ── 11. DMT ─────────────────────────────────────────────────────────
        df_dmt = pd.read_csv(archivos_recibidos["dmt"], sep=';', encoding='utf-8')
        lineas_eliminar = [
            'LINEA RZ-1','LINEA RZ-10','LINEA RZ-11',
            'LINEA RZ-2','LINEA RZ-3','LINEA RZ-6',
            'LINEA RZ-7','LINEA RZ-8'
        ]
        df_dmt = df_dmt[~df_dmt['nombre_linea'].isin(lineas_eliminar)]
        df_dmt['DMT'] = pd.to_numeric(
            df_dmt['DMT'].astype(str).str.replace(',', '.', regex=False),
            errors='coerce'
        ).round(2)
        df_dmt['Línea'] = pd.to_numeric(
            df_dmt['nombre_linea'].astype(str).str.extract(r'(\d+)')[0],
            errors='coerce'
        )
        df_dmt = df_dmt[df_dmt['Línea'].notna() & (df_dmt['Línea'] != 0)]
        df_dmt['Línea'] = df_dmt['Línea'].astype('int64')
        df_dmt = df_dmt[['Línea', 'DMT']].sort_values('Línea')

        # ── 12. MERGE FINAL ─────────────────────────────────────────────────
        if 205 in df_dem_op_parque['Línea'].values:
            df_dem_op_parque = df_dem_op_parque[df_dem_op_parque['Línea'] != 205]
        if 'VC_goef' in df_dem_op_parque.columns:
            df_dem_op_parque = df_dem_op_parque.drop(columns=['VC_goef'])

        df_indicadores = df_dem_op_parque.merge(df_dmt, on='Línea', how='left')
        df_indicadores['Factor de ocupación'] = (
            (df_indicadores['DMT'] * df_indicadores['TRX total mes']) /
            df_indicadores['Km mensuales']
        ) / 67
        df_indicadores['IPK'] = df_indicadores['TRX total mes'] / df_indicadores['Km mensuales']

        # Columnas ordenadas (solo las que existen)
        orden = [
            'Línea', 'JURISDICCION', 'Empresa',
            'PM Hab', 'Max Veh Calle',
            'Antiguedad media de flota', 'Motor Euro3', 'Motor Euro5',
            'Motor Eléctrico', 'Motor a GNC',
            'DMT', 'VC', 'Tiempo de marcha mensual (en hs)',
            'Km día hábil', 'Km mensuales', 'Factor de ocupación', 'IPK',
            'prom TRX dia habil', 'TRX total mes',
            'TRX OP tarjetas bancarias (física o virtual)',
            'TRX total NSSA',
            'TRX SUBE física', 'TRX app SUBE', 'TRX SUBE (física + app)',
            'TRX QR', 'TRX tarjetas bancarias (física o virtual)',
            'TRX multipago (QRs + tarjetas)',
            'TRX con boleto escolar primario', 'TRX con boleto escolar secundario',
            'TRX con ATS', 'TRX a partir de 2da integración (ITG)'
        ]
        cols_presentes = [c for c in orden if c in df_indicadores.columns]
        df_indicadores = df_indicadores[cols_presentes]

        # ── 13. EXPORTAR al Excel base ──────────────────────────────────────
        df = df_indicadores.copy().replace(np.nan, 0)
        nombre_hoja = f'{mes}-{anio}'

        ruta_base = archivos_recibidos["indicadores"]
        try:
            wb = load_workbook(ruta_base)
        except Exception:
            wb = Workbook()

        if nombre_hoja in wb.sheetnames:
            del wb[nombre_hoja]

        ws = wb.create_sheet(title=nombre_hoja)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        ws.freeze_panes = 'B2'

        # Estilos
        fill_header   = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        fill_first    = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        fill_dark     = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        align_cw      = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_c       = Alignment(horizontal='center', vertical='center')
        font_blue     = Font(bold=True, color="00008B")
        font_red      = Font(bold=True, color="FF0000")
        font_grey     = Font(bold=True, color="D9D9D9")
        font_bold     = Font(bold=True)
        cols_especiales = {
            'TRX SUBE (física + app)',
            'TRX multipago (QRs + tarjetas)',
            'TRX total NSSA'
        }

        for i, cell in enumerate(ws[1], start=1):
            if cell.value in cols_especiales:
                cell.fill = fill_dark
                cell.font = font_grey
            elif i == 1:
                cell.fill = fill_first
                cell.font = font_red
            else:
                cell.fill = fill_header
                cell.font = font_blue
            cell.alignment = align_cw

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.fill  = fill_first
                cell.font  = font_bold
                cell.alignment = align_c

        for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
            ws.column_dimensions[col[0].column_letter].width = 14
            for cell in col[1:]:
                cell.alignment = align_c

        if nombre_hoja in wb.sheetnames:
            wb.active = wb.index(wb[nombre_hoja])

        # Guardar en memoria y devolver
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        nombre_descarga = f"INDICADORES_completo_{mes}-{anio}.xlsx"
        return send_file(
            out,
            as_attachment=True,
            download_name=nombre_descarga,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500

    finally:
        # Limpiar archivos temporales
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"Servidor corriendo en puerto {port}")
    app.run(host="0.0.0.0", port=port, debug=False)
